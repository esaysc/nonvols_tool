import pandas as pd
import re
import shutil
from openpyxl import load_workbook
import copy

def clean_name(name):
    if pd.isna(name):
        return ""
    return re.sub(r'\s+', '', str(name))

def process():
    source_excel = '选民登记册.xlsx'
    template_excel = '人员名单-格 - 行.xlsx'
    output_excel = '人员名单-格 - 行_最终版_v2.xlsx'
    
    # 1. 复制模板文件
    shutil.copy(template_excel, output_excel)
    
    # 2. 读取选民数据
    df_source = pd.read_excel(source_excel)
    
    # 3. 加载新生成的 Excel
    wb = load_workbook(output_excel)
    template_ws = wb.active # 获取第一个工作表作为模板
    template_ws.title = "小组1" # 重命名第一个表
    
    # 获取模板样式（用于后续新表）
    # 注意：openpyxl 复制 sheet 比较复杂，我们采用循环创建并复制属性的方式
    
    for i, col in enumerate(df_source.columns):
        names_in_col = df_source[col].dropna().apply(clean_name).tolist()

        # 统计姓名出现次数并准备标记
        name_counts = {}
        for name in names_in_col:
            if name:
                name_counts[name] = name_counts.get(name, 0) + 1

        name_markers = {}
        current_indices = {}
        for name, count in name_counts.items():
            if count == 2:
                name_markers[name] = ["(大)", "(小)"]
            elif count >= 3:
                name_markers[name] = ["(大)", "(中)", "(小)"] + [""] * (count - 3)
            else:
                name_markers[name] = [""]
            current_indices[name] = 0

        marked_names = []
        for name in names_in_col:
            if name in name_markers:
                idx = current_indices[name]
                marker = name_markers[name][idx] if idx < len(name_markers[name]) else ""
                marked_names.append(f"{name}{marker}")
                current_indices[name] += 1
            else:
                marked_names.append(name)

        # 获取或创建工作表
        if i == 0:
            ws = template_ws
        else:
            # 复制模板工作表
            ws = wb.copy_worksheet(template_ws)
            ws.title = f"小组{i+1}"

        # 修改表头中的小组名称 (第二行 A2)
        ws['A2'].value = f"第{i+1}村民小组"

        # 清除模板中原有的姓名数据（从第三行开始）
        # 假设模板中姓名区域可能很大，我们先清除一定范围
        for row in ws.iter_rows(min_row=3):
            for cell in row:
                cell.value = None

        # 写入新姓名数据
        last_row = 3
        for idx, name in enumerate(marked_names):
            row_idx = (idx // 8) + 3
            col_idx = (idx % 8) + 1
            ws.cell(row=row_idx, column=col_idx, value=name)
            last_row = max(last_row, row_idx)

        # 动态设置打印区域，确保底部的线正好在数据结束的地方
        # 假设表格宽度是 A 到 H 列 (8列)
        ws.print_area = f'A1:H{last_row}'

    wb.save(output_excel)
    print(f"处理完成，结果已保存至: {output_excel}")

if __name__ == "__main__":
    process()
