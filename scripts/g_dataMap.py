from openpyxl import load_workbook

# ===================== 仅需修改这里的配置 =====================
target_village = "天宫"   # 目标村名
target_year = "2026年"    # 目标年份
file_path = "./data/2026年2季度贵平镇 蔬菜- .xlsx"  # 你的Excel路径
# =================================================================

# 加载Excel文件
wb = load_workbook(file_path, data_only=True)
ws = wb.active

# 遍历合并单元格，找到目标村表头
target_merge_cell = None
for merge_cell in ws.merged_cells.ranges:
    cell_value = ws[merge_cell.min_row][merge_cell.min_col - 1].value
    if cell_value and target_village in str(cell_value) and target_year in str(cell_value):
        target_merge_cell = merge_cell
        break

if not target_merge_cell:
    print(f"❌ 未找到 {target_village}{target_year} 的表头")
    exit()

# 自动定位列
area_col = target_merge_cell.min_col
yield_col = area_col + 1
code_col = 2
data_start_row = 8

# 生成 dataMap
data_map = {}
for row in range(data_start_row, ws.max_row + 1):
    code = ws.cell(row=row, column=code_col).value
    area = ws.cell(row=row, column=area_col).value
    yield_val = ws.cell(row=row, column=yield_col).value

    if code is None:
        continue
    
    code_str = f"{int(code):02d}"
    area_str = str(area) if area is not None else ""
    yield_str = str(yield_val) if yield_val is not None else ""

    data_map[code_str] = {"area": area_str, "yield": yield_str}

# ===================== 直接输出【完整可运行浏览器脚本】 =====================
print("="*80)
print("✅ 直接复制以下全部内容，到浏览器控制台运行即可！")
print("="*80)

print(r"""
{
  const dataMap = {""")

# 输出 dataMap
for code, data in data_map.items():
    print(f'    "{code}": {{ area: "{data["area"]}", yield: "{data["yield"]}" }},')

print(r"""
  };

  const rows = document.querySelectorAll("table tbody tr");
  let fillCount = 0;

  rows.forEach(row => {
    const tds = row.querySelectorAll("td");
    const code = tds[1]?.textContent.trim();
    const data = dataMap[code];

    if (!data) return;

    // 面积：第3列
    const areaInput = tds[2]?.querySelector("input");
    if (areaInput && data.area) {
      areaInput.value = data.area;
      areaInput.dispatchEvent(new Event("input"));
      areaInput.dispatchEvent(new Event("change"));
    }

    // 产量：第7列
    const yieldInput = tds[6]?.querySelector("input");
    if (yieldInput && data.yield) {
      yieldInput.value = data.yield;
      yieldInput.dispatchEvent(new Event("input"));
      yieldInput.dispatchEvent(new Event("change"));
    }

    fillCount++;
    console.log(`✅ 编码 ${code} 填写完成`);
  });

  console.log(`\n🎉 全部填写完成！共填充 ${fillCount} 行`);
}
""")