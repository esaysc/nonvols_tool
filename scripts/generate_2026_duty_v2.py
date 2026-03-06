import calendar
from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import Border


def generate_2026_duty():
    template_file = "2025中岗村全年值班表.xlsx"
    output_file = "2026中岗村全年值班表.xlsx"

    # 2026人员循环配置
    groups = [
        [
            ("王康", "18808208881"),
            ("苏琴", "17381138965"),
            ("程星", "18380225036"),
        ],
        [("夏桂群", "13219769699"), ("周召全", "15182225421")],
        [("熊燕", "15183300229"), ("唐晓兰", "13890333407")],
        [("王宇", "18328128986"), ("晏王琪", "13541225712")],
    ]

    wb = load_workbook(template_file)
    group_idx = 0
    weekdays_cn = ["星期一", "星期二", "星期三", "星期四", "星期五", "星期六", "星期日"]

    for month in range(1, 13):
        sheet_name = f"{month}月"
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]

        # 1. 修改标题
        if ws["A1"].value:
            ws["A1"].value = f"2026年{month}月值班表"

        # 2. 清除原有的数据行（从第3行开始）
        # 我们先记录第3行的样式作为模板
        template_cells = []
        for col in range(1, 5):
            template_cells.append(copy(ws.cell(row=3, column=col)))

        # 清除第3行及之后的所有内容和合并单元格（保留样式可能比较复杂，我们采取覆盖策略）
        # 简单起见，我们直接从第3行开始重新写入，并应用样式

        # 获取原有的合并单元格并移除（仅限第3行以后的）
        merges = list(ws.merged_cells.ranges)
        for merge in merges:
            if merge.min_row >= 3:
                ws.unmerge_cells(str(merge))

        # 清除第3行之后的数据
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=4):
            for cell in row:
                cell.value = None

        # 3. 写入2026年数据
        first_weekday, num_days = calendar.monthrange(2026, month)
        current_row = 3

        for day in range(1, num_days + 1):
            curr_date = f"2026-{month:02d}-{day:02d}"
            wd = (first_weekday + day - 1) % 7
            curr_week = weekdays_cn[wd]

            current_group = groups[group_idx]
            num_people = len(current_group)

            # 合并日期和星期
            if num_people > 1:
                ws.merge_cells(
                    start_row=current_row,
                    start_column=1,
                    end_row=current_row + num_people - 1,
                    end_column=1,
                )
                ws.merge_cells(
                    start_row=current_row,
                    start_column=2,
                    end_row=current_row + num_people - 1,
                    end_column=2,
                )

            # 填充数据并应用样式
            for i, (name, phone) in enumerate(current_group):
                row_idx = current_row + i

                # 日期列
                c1 = ws.cell(row=row_idx, column=1)
                if i == 0:
                    c1.value = curr_date
                c1.font = copy(template_cells[0].font)
                c1.border = copy(template_cells[0].border)
                c1.alignment = copy(template_cells[0].alignment)
                c1.number_format = template_cells[0].number_format

                # 星期列
                c2 = ws.cell(row=row_idx, column=2)
                if i == 0:
                    c2.value = curr_week
                c2.font = copy(template_cells[1].font)
                c2.border = copy(template_cells[1].border)
                c2.alignment = copy(template_cells[1].alignment)

                # 人员列
                c3 = ws.cell(row=row_idx, column=3)
                c3.value = name
                c3.font = copy(template_cells[2].font)
                c3.border = copy(template_cells[2].border)
                c3.alignment = copy(template_cells[2].alignment)

                # 电话列
                c4 = ws.cell(row=row_idx, column=4)
                c4.value = phone
                c4.font = copy(template_cells[3].font)
                c4.border = copy(template_cells[3].border)
                c4.alignment = copy(template_cells[3].alignment)

                # 设置行高
                ws.row_dimensions[row_idx].height = 20

            current_row += num_people
            group_idx = (group_idx + 1) % 4

        # 清除可能残留的旧行（如果2026年该月行数少于2025年）
        # 实际上上面已经清空了value，这里可以根据需要删除多余行
        if ws.max_row > current_row - 1:
            # 简单处理：确保后面的行没有边框和内容
            for r in range(current_row, ws.max_row + 1):
                for c in range(1, 5):
                    cell = ws.cell(row=r, column=c)
                    cell.value = None
                    cell.border = Border()

    wb.save(output_file)
    print(f"Successfully generated {output_file} using template.")


if __name__ == "__main__":
    generate_2026_duty()
